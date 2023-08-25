#!/usr/bin/env python3


##----------------------------------------------------------------------imports-----------------------------------------------------------------
import openpyxl
from openpyxl.styles import Font, colors
import sys
import numpy as np
import math as m

##------------------------------------------------------------------function defs---------------------------------------------------------------
def add_to_sheet(ind, a0, a1, a2, a3, a4, a5):
    ws['A{_}'.format(_ = ind + 5)] = a0
    ws['B{_}'.format(_ = ind + 5)] = a1
    ws['C{_}'.format(_ = ind + 5)] = a2
    ws['D{_}'.format(_ = ind + 5)] = a3
    ws['E{_}'.format(_ = ind + 5)] = a4
    if a5 == "Colliding":
        ws['F{_}'.format(_ = ind + 5)] = a5
        ws['F{_}'.format(_ = ind + 5)].font = Font(color="FF0000")
    else:
        ws['F{_}'.format(_ = ind + 5)] = a5
        ws['F{_}'.format(_ = ind + 5)].font = Font(color="00FF00")
    wb.save("../data_collection/5axis.xlsx")


##----------------------------------------------------------------------------------------------------------------------------------------------
if __name__ == '__main__':
    wb = openpyxl.load_workbook("../data_collection/5axis.xlsx")
    ws = wb.active

    ws['A2'] = float(sys.argv[1])
    ws['B2'] = float(sys.argv[2])
    ws['C2'] = float(sys.argv[3])
    ws['D2'] = float(sys.argv[4])
    ws['E2'] = float(sys.argv[5])
    ws['F2'] = float(sys.argv[6])
    ws['G2'] = float(sys.argv[7])
    ws['H2'] = float(sys.argv[8])


    wb.save("../data_collection/5axis.xlsx")