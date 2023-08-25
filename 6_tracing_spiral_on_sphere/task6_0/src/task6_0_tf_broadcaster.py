#!/usr/bin/env python

##---------------------------------------------------------------------Imports------------------------------------------------------------------

import rospy

import tf_conversions

import tf2_ros
import geometry_msgs.msg
import math as m
import numpy as np

import openpyxl


##------------------------------------------------------------Assignment of Variables-----------------------------------------------------------

##----------------------------------------------------------------------------------------------------------------------------------------------

i = 0
it = 0
forward_flag = True
##------------------------------------------------------------------Function Defs---------------------------------------------------------------

def x_z_from_pitch(thetap, radius_of_sphere, length_of_spindle, z_offset = 0.025):
    x = np.zeros(120)
    z = np.zeros(120)
    for _ in range(120):
        x[_] = (radius_of_sphere * m.sin(thetap[_])) + (length_of_spindle * m.sin(thetap[_]))
        z[_] = z_offset + radius_of_sphere + (radius_of_sphere * m.cos(thetap[_])) + (length_of_spindle * m.cos(thetap[_]))

    print(" x : {_}".format(_ = x))
    print(" z : {_}".format(_ = z))

    return x,z

def add_to_sheet():
    ws['A1'] = "X (mm)"
    ws['B1'] = "Y (mm)"
    ws['C1'] = "Z (mm)"
    ws['D1'] = "Pitch (radians)"
    ws['E1'] = "Pitch (degrees)"

    for _ in range(120):
        ws['A{_}'.format(_ = _ + 2)] = x[_] * 1000
        ws['B{_}'.format(_ = _ + 2)] = y[_] * 1000
        ws['C{_}'.format(_ = _ + 2)] = z[_] * 1000
        ws['D{_}'.format(_ = _ + 2)] = thetap[_]
        ws['E{_}'.format(_ = _ + 2)] = thetap[_] * 180 / m.pi

    # wb.save("../data_collection/5axis.xlsx")
    wb.save("../my_ws/src/task6_0/data_collection/5axis.xlsx")


# def circular_path_calculator(x , radius, x_c, y_c, positive_flag):
#     y = np.zeros(120)
#     # print(type(x))
#     x_new = np.linspace(0.1767766953,-0.1767766953,60)
#     x = np.concatenate((x,x_new),axis=0)
#     print(len(x))
#     for i in range(61):
#         y[i] = m.sqrt((radius)**2 - (x[i] - x_c)**2) + y_c
#     for i in range(61,120):
#         y[i] = -m.sqrt((radius)**2 - (x[i] - x_c)**2) + y_c

#     # if positive_flag:
#     #     y[it] = m.sqrt(radius**2 - (x[it] - x_c)**2) + y_c
#     # else:
#     #     y[it] = -m.sqrt(radius**2 - (x[it] - x_c)**2) + y_c
#     print("Y : {_}".format(_ = y))
#     return x,y

# def circular_path_polar(thetay, rad_sphere, link_length):
#     x = []
#     y = []
#     l = (rad_sphere/m.sqrt(2)) + (link_length/m.sqrt(2))
#     for _ in thetay:
#         x.append(l * m.cos(_))
#         y.append(l * m.sin(_))
#     print("x : \n{_}".format(_ = x))
#     print("y : \n{_}".format(_ = y))
#     return x,y

# # def circular_translation_from_x(x):
# #     for i,val in enumerate(x):
# #         if i > len(x) / 2:
# #             positive_flag_for_y = True
# #         if positive_flag_for_y :
# #             # y[i] = m.sqrt((radius)**2 - (x[i] - x_c)**2) + y_c

def handle_joint_pose(x, y, z,thetap, thetay):
    br = tf2_ros.TransformBroadcaster()

    t0 = geometry_msgs.msg.TransformStamped()
    t0.header.stamp = rospy.Time.now()
    t0.header.frame_id = "base_link"
    t0.child_frame_id = "x_axis_link"
    t0.transform.translation.x = x[it]    ##x[it]
    t0.transform.translation.y = 0
    t0.transform.translation.z = 0
    q = tf_conversions.transformations.quaternion_from_euler(0,0,0)
    t0.transform.rotation.x = q[0]
    t0.transform.rotation.y = q[1]
    t0.transform.rotation.z = q[2]
    t0.transform.rotation.w = q[3]
    br.sendTransform(t0)

    t2 = geometry_msgs.msg.TransformStamped()
    t2.header.stamp = rospy.Time.now()
    t2.header.frame_id = "base_link"
    t2.child_frame_id = "y_axis_link"
    t2.transform.translation.x = 0
    t2.transform.translation.y = y[it]  ##y[it]
    t2.transform.translation.z = 0
    q = tf_conversions.transformations.quaternion_from_euler(0,0,0)
    t2.transform.rotation.x = q[0]
    t2.transform.rotation.y = q[1]
    t2.transform.rotation.z = q[2]
    t2.transform.rotation.w = q[3]
    br.sendTransform(t2)

    t3 = geometry_msgs.msg.TransformStamped()
    t3.header.stamp = rospy.Time.now()
    t3.header.frame_id = "base_link"
    t3.child_frame_id = "object_link"
    t3.transform.translation.x = x[it]
    t3.transform.translation.y = y[it]
    t3.transform.translation.z = 0.05 + 0.025
    q = tf_conversions.transformations.quaternion_from_euler(0,0,thetay[i])
    t3.transform.rotation.x = q[0]
    t3.transform.rotation.y = q[1]
    t3.transform.rotation.z = q[2]
    t3.transform.rotation.w = q[3]
    br.sendTransform(t3)

    # Breaking some rules here, as even though this is only a prismatic joint , we are also giving it a yaw
    t4 = geometry_msgs.msg.TransformStamped()
    t4.header.stamp = rospy.Time.now()
    t4.header.frame_id = "base_link"
    t4.child_frame_id = "spindle_base_link"
    t4.transform.translation.x = 0
    t4.transform.translation.y = 0
    t4.transform.translation.z = z[it]  
    q = tf_conversions.transformations.quaternion_from_euler(0,0,3.1415)
    t4.transform.rotation.x = q[0]
    t4.transform.rotation.y = q[1]
    t4.transform.rotation.z = q[2]
    t4.transform.rotation.w = q[3]
    br.sendTransform(t4)

    t5 = geometry_msgs.msg.TransformStamped()
    t5.header.stamp = rospy.Time.now()
    t5.header.frame_id = "spindle_base_link"
    t5.child_frame_id = "spindle_nozzle_link"
    t5.transform.translation.x = 0
    t5.transform.translation.y = 0
    t5.transform.translation.z = 0
    q = tf_conversions.transformations.quaternion_from_euler(0,thetap[it] + 3.1415,0)
    t5.transform.rotation.x = q[0]
    t5.transform.rotation.y = q[1]
    t5.transform.rotation.z = q[2]
    t5.transform.rotation.w = q[3]
    br.sendTransform(t5)


def iterator_handler():
    global i
    global it

    if it == len(thetap) - 1:
        it = 0
    if i == len(thetay) - 1:
        i = 0
        it = it + 1
    else:
        i = i + 1


##----------------------------------------------------------------------Main Func---------------------------------------------------------------

if __name__ == '__main__':
    rospy.init_node('tf2_broadcaster')
    rate = rospy.Rate(120)

    wb = openpyxl.load_workbook("../my_ws/src/task6_0/data_collection/5axis.xlsx")                         ##Modifying the path for roslaunch 
    # wb = openpyxl.load_workbook("../data_collection/5axis.xlsx")
    ws = wb.active
    ws.title = "Sheet1"

    radius = 0.05

    # x_one = np.arange(-radius, radius, 0.001)
    # x_rev = np.flip(x_one)
    # np.delete(x_rev, 0)

    # x = np.concatenate((x_one, x_rev),axis = 0)
    # print(x)

    # x = np.linspace( -0.1767766953, 0.1767766953,60)
    # x,y = circular_path_calculator(x, 0.17677669530, 0, 0, forward_flag)
    # x = np.linspace(-0.300,0.300,60)
    # y = np.linspace(-0.300,0.300,60)
    # z = np.linspace(0.1 + 0.025 + 0.200,0.1 + 0.025 + 0.200,120)
    

    # x = np.zeros(120)
    y = np.zeros(120)
    # z = np.zeros(120)

    # z = z + 0.05 + 0.025 + (0.05 * m.cos(m.pi/4)) + (0.2 * m.cos(m.pi/4))

    # print(len(x))
    # print(len(y))
    # print(len(z))

    # print("x : {_}".format(_ = x))
    # print("y : {_}".format(_ = y))
    # print("z : {_}".format(_ = z))


    thetap = np.linspace(0, m.pi/4, 120)
    thetay = np.linspace(m.pi, -m.pi, 120)

    # thetap = np.zeros(120)
    # thetay = np.zeros(120)

    # x,y = circular_path_polar(thetay, 0.05, 0.2)
    x,z = x_z_from_pitch(thetap, radius_of_sphere=0.05, length_of_spindle=0.200, z_offset=0.025)

    add_to_sheet()

    while not rospy.is_shutdown():
        handle_joint_pose(x,y,z,thetap,thetay)
        iterator_handler()
        rate.sleep()
