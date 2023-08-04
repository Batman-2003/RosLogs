#!/usr/bin/env python

##---------------------------------------------------------------------imports------------------------------------------------------------------

import rospy

import tf_conversions

import tf2_ros
import geometry_msgs.msg
import math as m
import numpy as np

import sys

import openpyxl

##------------------------------------------------------------Assignment of Variables-----------------------------------------------------------

##----------------------------------------------------------------------------------------------------------------------------------------------

it = 0
forward_flag = True
##------------------------------------------------------------------Function Defs---------------------------------------------------------------
def read_parameters():
    x1 = float(ws['A2'].value)
    y1 = float(ws['B2'].value)
    x2 = float(ws['C2'].value)
    y2 = float(ws['D2'].value)
    z  = float(ws['E2'].value)
    p1 = float(ws['F2'].value)
    p2 = float(ws['G2'].value)
    p3 = float(ws['H2'].value)
    xo = float(ws['I2'].value)
    yo = float(ws['J2'].value)

    return [x1,y1,x2,y2,z,p1,p2,p3,xo,yo]

def path_calculator(x1,y1,x2,y2):
    slope = (y2 - y1) / (x2 - x1)
    c = -x1 + y1

    x = np.linspace(x1, x2, 60)
    y = slope * x + c
    return x,y

def handle_joint_pose():

    theta0 = m.atan(y[it]/x[it])

    L = m.sqrt(x[it]**2 + y[it]**2)
    K = m.sqrt(z**2 + L**2)

    alpha = m.acos((K/2) / 1.0)
    gamma = m.atan(z/L)

    theta1 = gamma + alpha
    theta2 = -2 * alpha

    br = tf2_ros.TransformBroadcaster()

    t0 = geometry_msgs.msg.TransformStamped()

    t0.header.stamp = rospy.Time.now()
    t0.header.frame_id = "base_link"
    t0.child_frame_id = "rear_arm_link"
    t0.transform.translation.x = 0
    t0.transform.translation.y = 0
    t0.transform.translation.z = 0.1
    q = tf_conversions.transformations.quaternion_from_euler(0, 0, theta0)
    t0.transform.rotation.x = q[0]
    t0.transform.rotation.y = q[1]
    t0.transform.rotation.z = q[2]
    t0.transform.rotation.w = q[3]
    br.sendTransform(t0)

    t1 = geometry_msgs.msg.TransformStamped()
    t1.header.stamp = rospy.Time.now()
    t1.header.frame_id = "rear_arm_link"
    t1.child_frame_id = "fore_arm_link"
    t1.transform.translation.x = 0.06
    t1.transform.translation.y = 0
    t1.transform.translation.z = 1.0
    q = tf_conversions.transformations.quaternion_from_euler(theta1 + 1.57, 0, 0)
    t1.transform.rotation.x = q[0]
    t1.transform.rotation.y = q[1]
    t1.transform.rotation.z = q[2]
    t1.transform.rotation.w = q[3]
    br.sendTransform(t1)

    t2 = geometry_msgs.msg.TransformStamped()
    t2.header.stamp = rospy.Time.now()
    t2.header.frame_id = "fore_arm_link"
    t2.child_frame_id = "wrist_arm_link"
    t2.transform.translation.x = -0.06
    t2.transform.translation.y = 0
    t2.transform.translation.z = -1.0
    q = tf_conversions.transformations.quaternion_from_euler(theta2, 0, 0)
    t2.transform.rotation.x = q[0]
    t2.transform.rotation.y = q[1]
    t2.transform.rotation.z = q[2]
    t2.transform.rotation.w = q[3]
    br.sendTransform(t2)


def direction_handler():
    global it
    global forward_flag

    if it == 60:
        forward_flag = False
        it = it - 1
    elif it == 0:
        forward_flag = True
        it = it + 1


def iterator_handler():
    global it
    global forward_flag

    if forward_flag:
        it = it + 1
    else:
        it = it - 1


##----------------------------------------------------------------------Main Func---------------------------------------------------------------
wb = openpyxl.load_workbook("../my_ws/src/collision_0/data_collection/test.xlsx")                         ##Modifying the path for roslaunch 
ws = wb.active
ws.title = "Sheet1"


if __name__ == '__main__':
    rospy.init_node('tf2_broadcaster')
    rate = rospy.Rate(10)

    paras = read_parameters()
    z = paras[4]
    x,y = path_calculator(paras[0], paras[1], paras[2], paras[3])

    while not rospy.is_shutdown():
        handle_joint_pose()
        iterator_handler()
        direction_handler()
        rate.sleep()
