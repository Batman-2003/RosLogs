#!/usr/bin/env python


##----------------------------------------------------------------------imports-----------------------------------------------------------------
import openpyxl
from openpyxl.styles import Font, colors
import sys
import numpy as np
import math as m

##------------------------------------------------------------------function defs---------------------------------------------------------------

def path_calculator(x1,y1,x2,y2):
    print("x1 : {x1}\ty1 : {y1}\tx2 : {x2}\ty2 : {y2}".format(x1 = x1, y1 = y1 , x2 = x2, y2 = y2))
    slope = (y2 - y1) / (x2 - x1)
    c = -x1 + y1

    x = np.linspace(x1, x2, 60)
    y = slope * x + c
    return x,y

def cylinder_collision_checking(x_space, y_space, RAD_OF_CYL, X_OFFSET, Y_OFFSET):
    collision_flag = False
    for ind, value in enumerate(zip(x_space,y_space)):
        if (value[0] - X_OFFSET)**2 + (value[1] - Y_OFFSET)**2 <= RAD_OF_CYL**2:
            collision_flag = True
            angle0, angle1, angle2 = getAngle(value[0], value[1])
            add_to_sheet(ind, angle0, angle1, angle2, value[0], value[1], "Colliding")
        else:
            angle0, angle1, angle2 = getAngle(value[0], value[1])
            add_to_sheet(ind, angle0, angle1, angle2, value[0], value[1], "All Clear")
    return collision_flag


def add_to_sheet(ind, a0, a1, a2, a3, a4, a5):
    ws['A{_}'.format(_ = ind + 5)] = a0
    ws['B{_}'.format(_ = ind + 5)] = a1
    ws['C{_}'.format(_ = ind + 5)] = a2
    ws['D{_}'.format(_ = ind + 5)] = a3
    ws['E{_}'.format(_ = ind + 5)] = a4
    if a5 == "Colliding":
        ws['F{_}'.format(_ = ind + 5)] = a5
        ws['F{_}'.format(_ = ind + 5)].font = Font(color="FF0000")
    else:
        ws['F{_}'.format(_ = ind + 5)] = a5
        ws['F{_}'.format(_ = ind + 5)].font = Font(color="00FF00")
    wb.save("../data_collection/test.xlsx")


def askForPermission(flag):
    if flag:
        reply = raw_input("There are collision on the given path , still want to proceed ? ( y for yes ) : ")
        if reply == "y":
            print("OK Proceeding anyways")
        else:
            sys.exit()

def getAngle(_x, _y):
    ang0 = m.tan(_y / _x)
    L = m.sqrt(_x**2 + _y**2)
    K = m.sqrt( L**2 +  z**2)
    alpha = m.acos((K/2) / 1.0)
    gamma = m.atan(K / z)
    ang1 = gamma - alpha
    ang2 = -2 * alpha
    return ang0, ang1, ang2

##----------------------------------------------------------------------------------------------------------------------------------------------

wb = openpyxl.load_workbook("../data_collection/test.xlsx")

ws = wb.active

x1 = ws['A2'] = float(sys.argv[1])
y1 = ws['B2'] = float(sys.argv[2])
x2 = ws['C2'] = float(sys.argv[3])
y2 = ws['D2'] = float(sys.argv[4])
z  = ws['E2'] = float(sys.argv[5])
p1 = ws['F2'] = float(sys.argv[6])
p2 = ws['G2'] = float(sys.argv[7])
p3 = ws['H2'] = float(sys.argv[8])
x0 = ws['I2'] = float(sys.argv[9])
y0 = ws['J2'] = float(sys.argv[10])

x,y = path_calculator(x1,y1,x2,y2)
safety_flag = cylinder_collision_checking(x, y, p1, x0, y0)
askForPermission(safety_flag)


wb.save("../data_collection/test.xlsx")